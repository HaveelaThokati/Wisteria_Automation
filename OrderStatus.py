import json
import openpyxl
from langchain.chains import LLMChain
from langchain_community.chat_models import AzureChatOpenAI
from langchain.prompts import ChatPromptTemplate, HumanMessagePromptTemplate, SystemMessagePromptTemplate


def initialize_llm():
    return AzureChatOpenAI(
        deployment_name="vassar-4o-mini",
        model_name="gpt-4o-mini",
        azure_endpoint="https://vassar-openai.openai.azure.com/",
        openai_api_key="DWDh8ZelTMEab6XAys0wo0d4p71l2i58QHe8GTiWOwYo5bHzkTisJQQJ99ALACYeBjFXJ3w3AAABACOGJVPB",
        openai_api_version="2024-02-15-preview",
        temperature=0.0
    )


EXTRACTION_RULES = {
"SYSTEM": """You are a detail extraction bot. Your task is to extract the order ID(s) and email address from the USER CURRENT MESSAGE, refer to HISTORY when necessary.
    FOLLOW THE INSTRUCTIONS GIVEN BELOW:
    1. Order ID:
        a. Extract the order ID from the USER CURRENT MESSAGE.
        b. If the USER CURRENT MESSAGE does not include an order ID but mentions or implies a connection to a previous order, extract the order ID from the HISTORY.
        c. If no order ID is mentioned return an empty list [] for the order ID.
        d. If multiple order IDs are mentioned in the USER CURRENT MESSAGE, return a list of all the order IDs.
        e. If the email mentions a tracking number, do not consider it as an order ID.

    2. Email address:
        a. Extract the valid email address explicitly mentioned in the USER CURRENT MESSAGE.
        b. If the USER CURRENT MESSAGE does not include an email address but mentions or implies a connection to HISTORY, extract the email address explicitly mentioned from the HISTORY.
        c. Do not infer or assume or format the email address based on the names in the USER CURRENT MESSAGE and HISTORY.
        d. If no email address is mentioned, return an empty string ("") for the email address.
    """,
    "CONTEXT": """
        HISTORY:
         {previous_conversations}
        USER CURRENT MESSAGE:
         {question}
    """,
    "DISPLAY": """Ensure that the output is in the following JSON format exactly as shown:
    {{
            "order_ids": [List of order id strings],
            "email_address": "[Extracted email address]"
    }}
    """,
    "REMEMBER": """
    - First, extract details from the USER CURRENT MESSAGE.
    - Do not provide phone number and names as information for email address. """

}


def extract_order_details(email_history, current_email):
    llm = initialize_llm()
    prompt = ChatPromptTemplate(
        messages=[
            SystemMessagePromptTemplate.from_template(EXTRACTION_RULES["SYSTEM"]),
            HumanMessagePromptTemplate.from_template(EXTRACTION_RULES["CONTEXT"]),
            SystemMessagePromptTemplate.from_template(EXTRACTION_RULES["DISPLAY"]),
            SystemMessagePromptTemplate.from_template(EXTRACTION_RULES["REMEMBER"])
        ]
    )
    llm_chain = LLMChain(llm=llm, prompt=prompt, verbose=True)
    return llm_chain.run({"previous_conversations": email_history, "question": current_email})


def process_excel_data(sheet):
    for row in range(2, sheet.max_row + 1):  # sheet.max_row + 1
        email_content = sheet[f'B{row}'].value
        if not email_content:
            sheet[f'C{row}'] = sheet[f'D{row}'] = "Invalid Email"
            continue

        if "Users_Current_Email:" not in email_content:
            sheet[f'C{row}'] = sheet[f'D{row}'] = "Invalid Email"
            continue

        split_content = email_content.split("Users_Current_Email:", 1)
        email_history = split_content[0].strip() if split_content[0] else ""
        current_email = split_content[1].strip()

        if current_email:
            output = extract_order_details(email_history, current_email)
            if output:
                parsed_output = json.loads(output)
                if isinstance(parsed_output, dict):
                    print(parsed_output)
                    order_ids = parsed_output.get("order_ids", [])
                    email_address = parsed_output.get("email_address", "")

                    print(f"Row {row}: Extracted Order IDs - {order_ids}")
                    print(f"Row {row} Content: {email_content}")

                    sheet[f'C{row}'] = ", ".join(map(str,order_ids)) if order_ids else "No Order ID"
                    sheet[f'D{row}'] = email_address if email_address else "No Email"
                else:
                    sheet[f'C{row}'] = sheet[f'D{row}'] = "Invalid JSON Output"
            else:
                sheet[f'C{row}'] = sheet[f'D{row}'] = "No Output"


def process_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    process_excel_data(sheet)
    updated_file_path = file_path.replace(".xlsx", "_updated.xlsx")
    workbook.save(updated_file_path)
    print("Updated file saved at:", updated_file_path)


if __name__ == "__main__":
    file_path = "C:/Users/Haveela/Downloads/ID_Extraction_Automation.xlsx"

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    print(f"Results written to {file_path}")
    process_excel_data(sheet)

    workbook.save(file_path)
