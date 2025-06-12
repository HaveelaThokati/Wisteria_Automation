import json
import openpyxl
import requests
import re

URL = "https://connectedenterprise.vassardigital.ai/wisteria/api/gorgias/intent_classification_chat_testing"
HEADERS = {"Content-Type": "application/json"}
EXCEL_FILE_PATH = "C:/Users/Haveela/Downloads/ChatAutomation.xlsx"


# Initializing workbook
def initialize_workbook(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    return workbook, sheet


# Classifying intent and structuring the json format
def classify_intent_chat(row, sheet):
    conversation = sheet[f'C{row}'].value
    parts = conversation.split("USER_LATEST_CHAT:")
    if len(parts) < 2:
        raise ValueError(f"Invalid conversation format in row {row}. Expected 'USER_LATEST_CHAT:' delimiter.")

    user_latest_message = re.sub(r'^(Human:|human:)\s*', '', parts[1].strip(), flags=re.IGNORECASE)

    chat_history = parts[0].strip()
    exchanges = [e.strip() for e in chat_history.split("**") if e.strip()]
    print(f"Exchanges: {exchanges}")

    ordered_conversation = []
    current_entry = {}

    for exchange in exchanges:
        if exchange.lower().startswith("human:"):
            if current_entry:
                ordered_conversation.append(current_entry)
            current_entry = {"human": re.sub(r'^human:\s*', '', exchange, flags=re.IGNORECASE).strip()}

        elif exchange.lower().startswith("ai:") or exchange.lower().startswith("ai :"):
            ai_message = re.sub(r'^ai\s*:\s*', '', exchange, flags=re.IGNORECASE).strip()
            if "human" in current_entry:
                if "AI" in current_entry:
                    current_entry["AI"] += "\n" + ai_message
                else:
                    current_entry["AI"] = ai_message
            else:
                if current_entry:
                    ordered_conversation.append(current_entry)
                current_entry = {"AI": ai_message}
    if current_entry:
        ordered_conversation.append(current_entry)

    return {
        "conversation": ordered_conversation,
        "query": user_latest_message
    }


# Seding request to payload
def send_request_and_process(payload):
    response = requests.post(URL, json=payload, headers=HEADERS)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error: {response.status_code}", response.text)
        return {}


# Main function
def main():
    workbook, sheet = initialize_workbook(EXCEL_FILE_PATH)
    pass_count = 0
    fail_count = 0

    for row in range(2, sheet.max_row + 1):
        if sheet[f'A{row}'].value is None:
            break
        print(f"Processing row {row}...")

        try:
            payload = classify_intent_chat(row, sheet)
            print("payload :: ", json.dumps(payload, indent=4, ensure_ascii=False))
            sheet[f'R{row}'] = json.dumps(payload, ensure_ascii=False)

            response_data = send_request_and_process(payload)
            result = response_data.get('result', {})
            intent_result = result.get('intent_classification_result', {})
            intent_list = intent_result.get("intent", [])

            classified_intents = []
            intent_string = ""
            for item in intent_list:
                score = float(item.get("similarity_score", 0))
                if score >= 0.4:
                    classified_intents.append(item.get("intent_name"))
                    intent_string += f"{item['intent_name']} ({item['similarity_score']})\n"

            intentbotlikely = intent_result.get("bot_likely_response", "")
            sheet[f'F{row}'] = intent_string.strip()
            sheet[f'H{row}'] = intent_result.get("reason", "")
            sheet[f'I{row}'] = intentbotlikely.strip()

            sub_intent_raw_list = result.get("sub_intents", [])
            classified_sub_intents = []
            sub_intent_string = ""
            sub_reason_string = ""
            subintentbotlikely = ""

            for sub_intent_obj in sub_intent_raw_list:
                for main_intent, sub_data in sub_intent_obj.items():
                    sub_intents = sub_data.get("intent", [])
                    reason = sub_data.get("reason", "")
                    sub_response = sub_data.get("bot_likely_response", "")
                    sub_intents_str = ""
                    for sub in sub_intents:
                        score = float(sub.get("similarity_score", 0))
                        if score >= 0.4:
                            classified_sub_intents.append(sub.get("intent_name"))
                            sub_intents_str += f"{sub['intent_name']} ({sub['similarity_score']})\n"
                    if sub_intents_str:
                        sub_intent_string += f"{main_intent}: {sub_intents_str.strip()}\n"
                        if reason:
                            sub_reason_string += f"{main_intent}: {reason}\n"
                        if sub_response:
                            subintentbotlikely += f"{main_intent}: {sub_response.strip()}\n"

            sheet[f'G{row}'] = sub_intent_string.strip()
            sheet[f'J{row}'] = sub_reason_string.strip()
            sheet[f'K{row}'] = subintentbotlikely.strip()

            expected_intents = sheet[f'D{row}'].value.strip() if sheet[f'D{row}'].value else ""
            expected_sub_intents = sheet[f'E{row}'].value.strip() if sheet[f'E{row}'].value else ""

            expected_intents_list = [intent.strip() for intent in expected_intents.split(",") if intent.strip()]
            expected_sub_intents_list = [intent.strip() for intent in expected_sub_intents.split(",") if intent.strip()]

            is_intents_matched = set(expected_intents_list) == set(classified_intents)
            is_sub_intents_matched = set(expected_sub_intents_list) == set(classified_sub_intents)

            sheet[f'L{row}'] = 'PASS' if is_intents_matched and is_sub_intents_matched else 'FAIL'

            if is_intents_matched and is_sub_intents_matched:
                pass_count += 1
            else:
                fail_count += 1

        except Exception as e:
            print(f"Error processing row {row}: {e}")
            for col in 'FGHIJK':
                sheet[f'{col}{row}'] = "ERROR"
            sheet[f'L{row}'] = 'FAIL'
            fail_count += 1

    total_test_cases = sum(1 for row in range(2, sheet.max_row + 1) if sheet[f'A{row}'].value is not None)
    pass_percentage = (pass_count / total_test_cases) * 100 if total_test_cases > 0 else 0
    fail_percentage = (fail_count / total_test_cases) * 100 if total_test_cases > 0 else 0

    sheet['M2'] = total_test_cases
    sheet['N2'] = pass_count
    sheet['O2'] = fail_count
    sheet['P2'] = f"{pass_percentage:.2f}%"
    sheet['Q2'] = f"{fail_percentage:.2f}%"

    # Writing results to excel
    workbook.save(EXCEL_FILE_PATH)
    print(f"Results written to {EXCEL_FILE_PATH}")


if __name__ == "__main__":
    main()
