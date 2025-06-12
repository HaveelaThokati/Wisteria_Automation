import json
import openpyxl
import requests

URL = "https://wisteria-uat.vassardigital.ai/wisteria/api/gorgias/intent_classification_testing"
HEADERS = {"Content-Type": "application/json"}
EXCEL_FILE_PATH = "C:/Users/Haveela/Downloads/IntentClassification_ProdIssues4.xlsx"


# Initializing workbook
def initialize_workbook(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    column_widths = {
        'A': 10,
        'B': 30,
        'C': 100,
        'D': 20,
        'E': 20,
        'F': 40,
        'G': 40,
        'H': 100,
        'I': 50,
        'J': 50,
        'K': 100,
        'L': 50,
        'M': 10,
        'N': 15,
        'O': 15,
        'P': 15,
        'Q': 15,
        'R': 15,
        'U': 70
    }

    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
    return workbook, sheet


# Extracting subject and body
def extract_subject_body(email_text):
    subject_split = email_text.split('Subject:', 1)
    body_split = subject_split[1].split('Body:', 1)
    subject = body_split[0].strip()
    body = body_split[1].strip()
    return {"subject": subject, "body": body}


# Classifying intent
def classify_intent(row, sheet):
    conversation = sheet[f'C{row}'].value
    parts = conversation.split("USER_LATEST_EMAIL:")
    email_history = parts[0].strip()
    user_latest_email = parts[1].strip()

    exchanges = email_history.split("**")
    ordered_conversation = [
        {"human": exchange.replace("User:", "")} if "User:" in exchange else
        {"ai": exchange.replace("Bot:", "")}
        for exchange in exchanges if exchange.strip()
    ]

    query = extract_subject_body(user_latest_email)

    return {
        "conversation": ordered_conversation,
        "query": query
    }


# Sending request to payload
def send_request_and_process(payload):
    response = requests.post(URL, json=payload, headers=HEADERS)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error: {response.status_code}", response.text)
        return {}


# Main function
def main():
    workbook, sheet = initialize_workbook(EXCEL_FILE_PATH)
    pass_count = 0
    fail_count = 0

    for row in range(2, sheet.max_row + 1):  # 2, sheet.max_row + 1
        if sheet[f'A{row}'].value is None:
            break
        print(f"Processing row {row}...")

        payload = classify_intent(row, sheet)
        sheet[f'U{row}'] = str(payload)

        try:
            print("payload :: ", json.dumps(payload, indent=4))
            response_data = send_request_and_process(payload)

            result = response_data.get('result', {})

            major_intent_result = result.get('major_intent', {})
            major_intent_value = major_intent_result.get("intent", [])

            if isinstance(major_intent_value, list):
                major_intents_classified_list = [
                    item for item in major_intent_value if item.get("similarity_score", 0) >= 0.4
                ]
                classified_intents = [item.get("intent") for item in major_intents_classified_list]
                sheet[f'F{row}'] = ", ".join(
                    f"{item['intent']} ({item['similarity_score']})" for item in major_intents_classified_list)
            else:
                classified_intents = [major_intent_value] if major_intent_value else []
                sheet[f'F{row}'] = major_intent_value if major_intent_value else ""

            sheet[f'G{row}'] = result.get("user_input", "")
            sheet[f'H{row}'] = major_intent_result.get("examples", "")
            sheet[f'I{row}'] = major_intent_result.get("reason", "")

            sub_intent_result = result.get('sub_intents', {})
            classified_sub_intents = []
            sub_intent_string, examples1, sub_intent_reason = "", "", ""

            if isinstance(major_intent_value, list):
                for major_intent in major_intents_classified_list:
                    intent = major_intent.get("intent")
                    intent_json = sub_intent_result.get(intent, {})
                    sub_intents_list = [
                        item for item in intent_json.get("sub_intents", []) if
                        item.get("similarity_score", 0) >= 0.4
                    ]
                    sub_intents_str = ", ".join(
                        f"{item['sub_intent']} ({item['similarity_score']})" for item in sub_intents_list)

                    classified_sub_intents.extend([item.get("sub_intent") for item in sub_intents_list])
                    sub_intent_string += f"{intent}: {sub_intents_str}\n" if sub_intents_str else ""
                    examples1 += f"{intent}: {intent_json.get('examples', '')}\n" if intent_json.get(
                        'examples') else ""
                    sub_intent_reason += f"{intent}: {intent_json.get('reason', '')}\n" if intent_json.get(
                        'reason') else ""

            sheet[f'J{row}'] = sub_intent_string.strip()
            sheet[f'K{row}'] = examples1.strip()
            sheet[f'L{row}'] = sub_intent_reason.strip()

            expected_intents = sheet[f'D{row}'].value.strip() if sheet[f'D{row}'].value else ""
            expected_sub_intents = sheet[f'E{row}'].value.strip() if sheet[f'E{row}'].value else ""

            expected_intents_list = [intent.strip() for intent in expected_intents.split(",") if intent.strip()]
            expected_sub_intents_list = [intent.strip() for intent in expected_sub_intents.split(",") if intent.strip()]

            is_intents_matched = set(expected_intents_list) == set(classified_intents)
            is_sub_intents_matched = set(expected_sub_intents_list) == set(classified_sub_intents)

            if is_intents_matched and is_sub_intents_matched:
                sheet[f'M{row}'] = 'PASS'
                pass_count += 1
            else:
                sheet[f'M{row}'] = 'FAIL'
                fail_count += 1

        except Exception as e:
            print(f"Error processing row {row}: {e}")
            for col in 'FGHIJKL':
                sheet[f'{col}{row}'] = "ERROR"

    total_test_cases = sum(1 for row in range(2, sheet.max_row + 1) if sheet[f'A{row}'].value is not None)
    pass_percentage = (pass_count / total_test_cases) * 100 if total_test_cases > 0 else 0
    fail_percentage = (fail_count / total_test_cases) * 100 if total_test_cases > 0 else 0

    sheet[f'N2'] = total_test_cases
    sheet[f'O2'] = pass_count
    sheet[f'P2'] = fail_count
    sheet[f'Q2'] = f"{pass_percentage:.2f}%"
    sheet[f'R2'] = f"{fail_percentage:.2f}%"

    workbook.save(EXCEL_FILE_PATH)
    print(f"Results written to {EXCEL_FILE_PATH}")


if __name__ == "__main__":
    main()
