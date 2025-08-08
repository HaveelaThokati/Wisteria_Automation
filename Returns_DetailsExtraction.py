import json
import requests
from openpyxl import load_workbook

# API URL
API_URL = "https://connectedenterprise.vassardigital.ai/wisteria/api/gorgias/details_extraction_testing"


def send_to_api(input_json):
    headers = {"Content-Type": "application/json"}
    try:
        response = requests.post(API_URL, headers=headers, json=input_json)
        return response.status_code, response.json()
    except Exception as e:
        return None, {"error": str(e)}


def process_excel(file_path):
    wb = load_workbook(filename=file_path)
    ws = wb.active

    ws.cell(row=1, column=3).value = "Body Order IDs"
    ws.cell(row=1, column=4).value = "Body Email"
    ws.cell(row=1, column=5).value = "Subject Order IDs"
    ws.cell(row=1, column=6).value = "Status Code"
    ws.cell(row=1, column=7).value = "Remarks"
    ws.cell(row=1, column=8).value = "Is Refused Shipment"
    ws.cell(row=1, column=9).value = "Item Names"

    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=2).value

        cell_str = str(cell_value) if cell_value is not None else ""

        if not cell_str.strip():
            ws.cell(row=row, column=7).value = "Empty cell"
            continue

        try:
            input_json = json.loads(cell_str.strip())
        except json.JSONDecodeError as e:
            ws.cell(row=row, column=6).value = None
            ws.cell(row=row, column=7).value = f"Invalid JSON: {e}"
            continue

        # Send request to API
        status_code_http, response = send_to_api(input_json)

        body_order_ids = []
        body_email = ""
        subject_order_ids = []
        is_refused_shipment = ""
        item_names = []
        remark = ""


        if (
                isinstance(response, dict)
                and response.get("status") is True
                and "result" in response
        ):
            result = response.get("result", {})
            body = result.get("body_details_extracted", {})
            subject = result.get("subject_details_extracted", {})

            body_order_ids = body.get("order_ids", [])
            body_email = body.get("email_address", "")
            item_names = body.get("item_names", [])

            subject_order_ids = subject.get("order_ids", [])

            status_code_api = response.get("code", None) or status_code_http
            remark = "Success"
        else:
            status_code_api = status_code_http
            remark = response.get("error", "Invalid API response")

        ws.cell(row=row, column=3).value = ", ".join(map(str, body_order_ids)) if body_order_ids else ""
        ws.cell(row=row, column=4).value = body_email or ""
        ws.cell(row=row, column=5).value = ", ".join(map(str, subject_order_ids)) if subject_order_ids else ""
        ws.cell(row=row, column=6).value = status_code_api
        ws.cell(row=row, column=7).value = remark
        ws.cell(row=row, column=8).value = ""
        ws.cell(row=row, column=9).value = ", ".join(map(str, item_names)) if item_names else ""

        print(f"Row {row} processed — Status: {status_code_api}, Remark: {remark}")

    # Saving updated Excel
    wb.save(file_path)
    print("✅ Excel updated and saved:", file_path)


excel_file = "C:/Users/Haveela/Downloads/ReturnsExtraction.xlsx"
process_excel(excel_file)
