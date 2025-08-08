import json
import openpyxl
from langchain.chains import LLMChain
from langchain_community.chat_models import AzureChatOpenAI
from langchain.prompts import ChatPromptTemplate, HumanMessagePromptTemplate, SystemMessagePromptTemplate


def initialize_llm():
    return AzureChatOpenAI(
        deployment_name="vassar-4o-mini",
        model_name="gpt-4o-mini",
        azure_endpoint="https://vassar-openai.openai.azure.com/",
        openai_api_key="DWDh8ZelTMEab6XAys0wo0d4p71l2i58QHe8GTiWOwYo5bHzkTisJQQJ99ALACYeBjFXJ3w3AAABACOGJVPB",
        openai_api_version="2024-02-15-preview",
        temperature=0.0
    )


EXTRACTION_RULES = {
    "SYSTEM":"""You are a detail extraction bot. You are tasked with extracting details regarding damages from the provided CURRENT EMAIL content and if necessary, refer to EMAIL HISTORY for additional context.
    Sometimes user may answer BOT question directly with required information without any description or what the information refers to. Please use BOT question or EMAIL HISTORY to understand USER message.
    EMAIL HISTORY contains the conversation in chronological order, from the oldest to the most recent.
    While extracting order id and email address from EMAIL HISTORY consider USER message and BOT response to understand the context.
    The following instructions guide how to handle various scenarios:
    1. Order ID:
        a. Extract the order ID from the CURRENT EMAIL.
        b. If the user refers to a specific previous order without mentioning a new order ID, and there is a clear reference to the EMAIL HISTORY, extract the order ID from that prior conversation.
        c. If no order ID is mentioned return an empty list [] for the order ID.
        d. If multiple order IDs are mentioned, return a list of all the order IDs.
        e. If order id is found both in email body and subject, prioritize the order id in email body.
        f. If the email mentions a tracking number, do not consider it as an order ID.
        g. If the email mentions a product identifier (e.g., Product #) and it is in the context of a damage inquiry, treat it as an order number rather than an item.
    2. List of Item Names:
        a. Extract the item names from the CURRENT EMAIL and ensure it is a name (e.g., 'Barley Twist Coffee Table').
        b. When the user mentions a item by its position from a list in the EMAIL HISTORY, locate the list and extract the corresponding item details.
    3. Refused shipment:
        a. Determine if the shipment was refused due to damages. A refused shipment occurs when the product arrives damaged, and the customer rejects the delivery, sending it back to the vendor. However, if the customer accepted the delivery and later reported the damage, this does not qualify as a refused shipment.
        b. If the email content does not explicitly state that the shipment was refused at delivery due to damages, label the status as "UNIDENTIFIED" rather than assuming it was refused.
    4. Email address:
        a. Extract the email address from the CURRENT EMAIL.
        b. If the CURRENT EMAIL does not include an email address but mentions or implies a connection to a previous message, extract the email address from the EMAIL HISTORY.
        c. If the email address is not mentioned, return an empty string "" for the email address.
    """,
    "CONTEXT": """
    EMAIL HISTORY:
     {previous_conversations}
    CURRENT EMAIL:
     {question}
    """,
    "DISPLAY": """Ensure that the output is in the following JSON format exactly as shown:
        {{
              "order_ids": [List of order id strings],
              "email_address": "[Email address mentioned in the email content]",
              "is_refused_shipment": [Indicates whether the shipment was refused by the recipient. Using Boolean 'True' for refusal and 'False' otherwise],
              "item_names": [List the names of the items or products mentioned in the email content]
        }}""",
    "REMEMBER":"""First, extract details from the current email. Check the email history in both user messages and bot response."""
}





def extract_order_details(email_history, current_email):
    llm = initialize_llm()
    prompt = ChatPromptTemplate(
        messages=[
            SystemMessagePromptTemplate.from_template(EXTRACTION_RULES["SYSTEM"]),
            HumanMessagePromptTemplate.from_template(EXTRACTION_RULES["CONTEXT"]),
            SystemMessagePromptTemplate.from_template(EXTRACTION_RULES["DISPLAY"]),
        ]
    )
    llm_chain = LLMChain(llm=llm, prompt=prompt, verbose=True)
    return llm_chain.run({"previous_conversations": email_history, "question": current_email})


def process_excel_data(sheet):
    for row in range(2, sheet.max_row + 1):
        email_content = sheet[f'B{row}'].value
        if not email_content:
            continue

        split_content = email_content.split("Users_Current_Email:")
        email_history = split_content[0].strip() if len(split_content) > 0 else ""
        current_email = split_content[1].strip() if len(split_content) > 1 else ""

        if email_history and current_email:
            output = extract_order_details(email_history, current_email)
            if output:
                parsed_output = json.loads(output)
                order_ids = parsed_output.get("order_ids", [])
                email_address = parsed_output.get("email_address", "")

                # Print extracted Order ID(s) to the console
                print(f"Row {row}: Extracted Order IDs - {order_ids}")

                sheet[f'C{row}'] = ", ".join(order_ids)
                sheet[f'D{row}'] = email_address
            else:
                sheet[f'C{row}'] = sheet[f'D{row}'] = "No Output"
        else:
            sheet[f'C{row}'] = sheet[f'D{row}'] = "Invalid Email"



def process_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    process_excel_data(sheet)
    updated_file_path = file_path.replace(".xlsx", "_updated.xlsx")
    workbook.save(updated_file_path)
    print("Updated file saved at:", updated_file_path)


if __name__ == "__main__":
    file_path = "C:/Users/Haveela/Downloads/DamagesID_Extraction_Automation.xlsx"

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    process_excel_data(sheet)
    workbook.save(file_path)
    print(f"Results written to {file_path}")
