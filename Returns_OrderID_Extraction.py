import json
import openpyxl
from langchain.chains import LLMChain
from langchain_community.chat_models import AzureChatOpenAI
from langchain.prompts import ChatPromptTemplate, HumanMessagePromptTemplate, SystemMessagePromptTemplate


def initialize_llm():
    return AzureChatOpenAI(
        deployment_name="vassar-4o-mini",
        model_name="gpt-4o-mini",
        azure_endpoint="https://vassar-openai.openai.azure.com/",
        openai_api_key="DWDh8ZelTMEab6XAys0wo0d4p71l2i58QHe8GTiWOwYo5bHzkTisJQQJ99ALACYeBjFXJ3w3AAABACOGJVPB",
        openai_api_version="2024-02-15-preview",
        temperature=0.0
    )


EXTRACTION_RULES = {
        "SYSTEM":"""You are a detail extraction bot. You are tasked with extracting details regarding returns from the provided CURRENT EMAIL and if necessary, refer to EMAIL HISTORY for additional context.
        Sometimes user may answer BOT question directly with required information without any description or what the information refers to. Please use BOT question or EMAIL HISTORY to understand USER message.
        EMAIL HISTORY contains the conversation in chronological order, from the oldest to the most recent.
        While extracting order id and email address from EMAIL HISTORY consider USER message and BOT response to understand the context.

        The following instructions guide how to handle various scenarios:
        1. Order ID:
            a. Extract the order ID from the CURRENT EMAIL.
            b. If the user refers to a specific previous order without mentioning a new order ID, and there is a clear reference to the EMAIL HISTORY, extract the order ID from that prior conversation.
            c. If the user uses phrases like "another Order" or "an order I placed" without giving an explicit order ID, return an empty list [] for the order ID.
            d. If no order ID is mentioned return an empty list [] for the order ID.
            e. If multiple order IDs are mentioned, return a list of all the order IDs.
            f. If order id is found both in email body and subject, prioritize the order id in email body.
            g. If the email mentions a product identifier (e.g., Product #) and it is in the context of a return inquiry, treat it as an order number rather than an item.
            h. If the email mentions a tracking number, do not consider it as an order ID.

        2. List of Item Names:
            a. If the user specifies returning all items, extract all items from the EMAIL HISTORY related to the referenced order ID.
            b. If the user specifies returning specific items, extract the item names from the current email.
            c. When the user mentions a item by its position from a list in the EMAIL HISTORY, locate the list and extract the corresponding item details.
            d. If the user explicitly states that no items will be returned, return an empty list [].
            e. If the user does not mention any specific items and does not imply returning all items, return an empty list [].

        3. Email address:
            a. Extract the email address from the CURRENT EMAIL.
            b. If the CURRENT EMAIL does not include an email address but mentions or implies a connection to a previous message, extract the email address from the EMAIL HISTORY.
            c. If the email address is not mentioned, return an empty string "" for the email address.
        """ ,
        "CONTEXT": """
        EMAIL HISTORY:
         {previous_conversations}
        CURRENT EMAIL:
         {question}
        """,
        "DISPLAY": """Ensure that the output is in the following JSON format exactly as shown:
        {{
            "order_ids": [List of order id strings],
            "email_address": "[Email address mentioned in the email content]",
            "item_names": [List the names of the items or products mentioned in the email content]
        }}""",
        "REMEMBER": """First, extract details from the current email. Check the email history in both user messages and bot response."""
}





def extract_order_details(email_history, current_email):
    llm = initialize_llm()
    prompt = ChatPromptTemplate(
        messages=[
            SystemMessagePromptTemplate.from_template(EXTRACTION_RULES["SYSTEM"]),
            HumanMessagePromptTemplate.from_template(EXTRACTION_RULES["CONTEXT"]),
            SystemMessagePromptTemplate.from_template(EXTRACTION_RULES["DISPLAY"]),
        ]
    )
    llm_chain = LLMChain(llm=llm, prompt=prompt, verbose=True)
    return llm_chain.run({"previous_conversations": email_history, "question": current_email})


def process_excel_data(sheet):
    for row in range(2, sheet.max_row + 1):
        email_content = sheet[f'B{row}'].value
        if not email_content:
            continue

        split_content = email_content.split("Users_Current_Email:")
        email_history = split_content[0].strip() if len(split_content) > 0 else ""
        current_email = split_content[1].strip() if len(split_content) > 1 else ""

        if email_history and current_email:
            output = extract_order_details(email_history, current_email)
            if output:
                parsed_output = json.loads(output)
                order_ids = parsed_output.get("order_ids", [])
                email_address = parsed_output.get("email_address", "")

                # Print extracted Order ID(s) to the console
                print(f"Row {row}: Extracted Order IDs - {order_ids}")

                sheet[f'C{row}'] = ", ".join(order_ids)
                sheet[f'D{row}'] = email_address
            else:
                sheet[f'C{row}'] = sheet[f'D{row}'] = "No Output"
        else:
            sheet[f'C{row}'] = sheet[f'D{row}'] = "Invalid Email"



def process_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    process_excel_data(sheet)
    updated_file_path = file_path.replace(".xlsx", "_updated.xlsx")
    workbook.save(updated_file_path)
    print("Updated file saved at:", updated_file_path)


if __name__ == "__main__":
    file_path = "C:/Users/Haveela/Downloads/DamagesID_Extraction_Automation.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    process_excel_data(sheet)
    workbook.save(file_path)
    print(f"Results written to {file_path}")
