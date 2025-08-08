import json
import requests
from openpyxl import load_workbook

# API endpoint
API_URL = "https://connectedenterprise.vassardigital.ai/wisteria/api/gorgias/details_extraction_testing"

def send_to_api(input_json):
    headers = {"Content-Type": "application/json"}
    try:
        response = requests.post(API_URL, headers=headers, json=input_json)
        return response.status_code, response.json()
    except Exception as e:
        return None, {"error": str(e)}

def process_excel(file_path):
    wb = load_workbook(filename=file_path)
    ws = wb.active

    # Add headers if not already present
    ws.cell(row=1, column=3).value = "Body Product Names"
    ws.cell(row=1, column=4).value = "Body SKU IDs"
    ws.cell(row=1, column=5).value = "Subject Product Names"
    ws.cell(row=1, column=6).value = "Subject SKU IDs"
    ws.cell(row=1, column=7).value = "Status Code"
    ws.cell(row=1, column=8).value = "Remarks"

    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=2).value

        if not cell_value:
            ws.cell(row=row, column=8).value = "Empty cell"
            continue

        try:
            input_json = json.loads(cell_value.strip())
        except json.JSONDecodeError as e:
            ws.cell(row=row, column=7).value = None
            ws.cell(row=row, column=8).value = f"Invalid JSON: {e}"
            continue

        # Send request
        status_code, response = send_to_api(input_json)

        # Initialize defaults
        body_product_names = []
        body_sku_ids = []
        subject_product_names = []
        subject_sku_ids = []

        if isinstance(response, dict) and "result" in response:
            result = response.get("result", {})
            body = result.get("body_details_extracted", {})
            subject = result.get("subject_details_extracted", {})

            body_product_names = body.get("product_names", [])
            body_sku_ids = body.get("sku_ids", [])
            subject_product_names = subject.get("product_names", [])
            subject_sku_ids = subject.get("sku_ids", [])

            # Mark as Fail if all extracted lists are empty
            if not body_product_names and not body_sku_ids and not subject_product_names and not subject_sku_ids:
                remark = "Fail"
            else:
                remark = "Success"
        else:
            remark = response.get("error", "Invalid API response")

        # Write results to Excel
        ws.cell(row=row, column=3).value = ", ".join(body_product_names)
        ws.cell(row=row, column=4).value = ", ".join(body_sku_ids)
        ws.cell(row=row, column=5).value = ", ".join(subject_product_names)
        ws.cell(row=row, column=6).value = ", ".join(subject_sku_ids)
        ws.cell(row=row, column=7).value = status_code
        ws.cell(row=row, column=8).value = remark

        print(f"Row {row} processed — Status: {status_code}, Remark: {remark}")

    wb.save(file_path)
    print("✅ Excel updated and saved:", file_path)

# Run the function on your Excel file
excel_file = "C:/Users/Haveela/Downloads/ID_Extraction_Updated.xlsx"
process_excel(excel_file)
