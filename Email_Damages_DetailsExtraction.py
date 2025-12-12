import json
import requests
from openpyxl import load_workbook

# API endpoint
API_URL = "https://connectedenterprise.vassardigital.ai/wisteria/api/gorgias/details_extraction_testing"

def send_to_api(input_json):
    headers = {"Content-Type": "application/json"}
    try:
        response = requests.post(API_URL, headers=headers, json=input_json)
        return response.status_code, response.json()
    except Exception as e:
        return None, {"error": str(e)}

def process_excel(file_path):
    wb = load_workbook(filename=file_path)
    ws = wb.active

    # Add headers if not already present
    ws.cell(row=1, column=3).value = "Body Order IDs"
    ws.cell(row=1, column=4).value = "Body Email"
    ws.cell(row=1, column=5).value = "Subject Order IDs"
    ws.cell(row=1, column=6).value = "Status Code"
    ws.cell(row=1, column=7).value = "Remarks"
    ws.cell(row=1, column=8).value = "Is Refused Shipment"
    ws.cell(row=1, column=9).value = "Item Names"

    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=2).value

        # Convert to string safely and strip whitespace
        cell_str = str(cell_value) if cell_value is not None else ""

        if not cell_str.strip():
            ws.cell(row=row, column=7).value = "Empty cell"
            continue

        try:
            input_json = json.loads(cell_str.strip())
        except json.JSONDecodeError as e:
            ws.cell(row=row, column=6).value = None
            ws.cell(row=row, column=7).value = f"Invalid JSON: {e}"
            continue

        # Send request
        status_code_http, response = send_to_api(input_json)

        # Initialize defaults
        body_order_ids = []
        body_email = ""
        subject_order_ids = []
        is_refused_shipment = ""
        item_names = []
        remark = ""

        # Extract data from nested "result" structure
        if isinstance(response, dict) and response.get("status") is True and "result" in response:
            result = response.get("result", {})
            body = result.get("body_details_extracted", {})
            subject = result.get("subject_details_extracted", {})

            body_order_ids = body.get("order_ids", [])
            body_email = body.get("email_address", "")
            subject_order_ids = subject.get("order_ids", [])

            # New fields:
            is_refused_shipment = body.get("is_refused_shipment", "")
            item_names = body.get("item_names", [])

            # Use API 'code' field as status code if present
            status_code_api = response.get("code", status_code_http)

            remark = "Success"
        else:
            status_code_api = status_code_http
            remark = response.get("error", "Invalid API response")

        # Write results to Excel
        ws.cell(row=row, column=3).value = ", ".join(body_order_ids)
        ws.cell(row=row, column=4).value = body_email
        ws.cell(row=row, column=5).value = ", ".join(subject_order_ids)
        ws.cell(row=row, column=6).value = status_code_api
        ws.cell(row=row, column=7).value = remark

        # Additional columns for new fields
        ws.cell(row=row, column=8).value = is_refused_shipment
        ws.cell(row=row, column=9).value = ", ".join(item_names)

        print(f"Row {row} processed — Status: {status_code_api}, Remark: {remark}")

    # Save updated Excel
    wb.save(file_path)
    print(" Excel updated and saved:", file_path)

# Run the function on your Excel file
excel_file = "C:/Users/Haveela/Downloads/damageExtraction.xlsx"
process_excel(excel_file)

